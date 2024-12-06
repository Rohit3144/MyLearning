import os
import sys
import pytest
from unittest import mock
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    import xlwings as xw

    excel_available = True
except ImportError:
    excel_available = False
from openpyxl import Workbook


# Add the src directory to the Python path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../src")))

import dixon_graph_generator_exp  # type: ignore
from dixon_graph_generator_exp import (  # type: ignore
    create_blank_excel,
    populate_data,
    add_chart,
    display_chart,
)


@pytest.fixture(scope="module")
def test_file_path():
    """Fixture to provide the file path for the test Excel file."""
    file_path = os.path.join(os.getcwd(), "test_line_chart.xlsx")
    yield file_path
    # Cleanup after test execution
    if os.path.exists(file_path):
        os.remove(file_path)


@pytest.fixture
def test_data_dict():
    """Fixture to provide test data dictionary."""
    return {
        "sheet_num": 0,  # First sheet in the workbook
        "sheet_name": "Bitrate",
        "x_axis": "Bitrate",
        "first_header": "Zyxel-2.4_Bitrate",
        "sec_header": "DXB-2.4_Bitrate",
        "x_axis_value": [50, 80, 120, 200, 300, 400, 500, 700, 800, 1000],
        "first_header_value": [0, 6.5, 23, 22, 29, 53, 40, 57, 43, 45],
        "sec_header_value": [0, 5.4, 13, 32, 39, 42, 42, 51, 48, 51],
    }


def test_create_blank_excel(test_file_path):
    """Test if a blank Excel file is created."""
    if os.path.exists(test_file_path):
        os.remove(test_file_path)  # Ensure the file doesn't exist initially

    create_blank_excel(test_file_path)

    # Assert if the file has been created
    assert os.path.exists(test_file_path), "The Excel file was not created."


def test_populate_data(test_file_path, test_data_dict):
    """Test populate_data() using both xlwings and openpyxl."""

    # Call the populate_data() function
    populate_data(test_file_path, test_data_dict)

    # Assert that the file is created
    assert os.path.exists(test_file_path), "Excel file was not created."

    # If running on Windows with Excel available, use xlwings
    if sys.platform == "win32" and excel_available:
        # Test xlwings behavior
        app = xw.App(visible=False)
        try:
            workbook = app.books.open(test_file_path)
            sheet = workbook.sheets[test_data_dict["sheet_name"]]

            # Check if headers and data are correct with xlwings
            assert (
                sheet.range("A1").value == test_data_dict["x_axis"]
            ), "X-Axis header mismatch with xlwings"
            assert (
                sheet.range("B1").value == test_data_dict["first_header"]
            ), "First header mismatch with xlwings"
            assert (
                sheet.range("C1").value == test_data_dict["sec_header"]
            ), "Second header mismatch with xlwings"
            assert (
                sheet.range("A2").value == test_data_dict["x_axis_value"][0]
            ), "X-Axis data mismatch with xlwings"
            assert (
                sheet.range("B2").value == test_data_dict["first_header_value"][0]
            ), "First header data mismatch with xlwings"
            assert (
                sheet.range("C2").value == test_data_dict["sec_header_value"][0]
            ), "Second header data mismatch with xlwings"

        finally:
            app.quit()

    else:
        # Test openpyxl behavior
        wb = load_workbook(test_file_path)
        sheet = wb[test_data_dict["sheet_name"]]

        # Check if headers and data are correct with openpyxl
        assert (
            sheet["A1"].value == test_data_dict["x_axis"]
        ), "X-Axis header mismatch with openpyxl"
        assert (
            sheet["B1"].value == test_data_dict["first_header"]
        ), "First header mismatch with openpyxl"
        assert (
            sheet["C1"].value == test_data_dict["sec_header"]
        ), "Second header mismatch with openpyxl"
        assert (
            sheet["A2"].value == test_data_dict["x_axis_value"][0]
        ), "X-Axis data mismatch with openpyxl"
        assert (
            sheet["B2"].value == test_data_dict["first_header_value"][0]
        ), "First header data mismatch with openpyxl"
        assert (
            sheet["C2"].value == test_data_dict["sec_header_value"][0]
        ), "Second header data mismatch with openpyxl"


def test_add_chart(test_file_path):
    """Test if a chart is added to the Excel file."""
    chart_title = "Bitrate Comparison"
    add_chart(test_file_path, 0, chart_title)

    if sys.platform == "win32" and excel_available:
        app = xw.App(visible=False)
        try:
            workbook = app.books.open(test_file_path)
            sheet = workbook.sheets[0]

            # Check if chart exists
            assert len(sheet.charts) > 0, "Chart was not added to the sheet."

            # Verify chart title
            chart = sheet.charts[0]
            assert chart.name == chart_title, "Chart title is incorrect."

        finally:
            workbook.close()
            app.quit()
    else:
        # Check for chart in openpyxl
        wb = load_workbook(test_file_path)
        sheet = wb.active

        # Verify that the chart is created
        charts = sheet._charts
        assert len(charts) > 0, "Chart was not added to the sheet."

        # Debug: Print actual title for verification
        print(f"Expected Title: {chart_title}")
        actual_title = charts[0].title
        print(f"Actual Title: {actual_title}")

        # Check chart title #TODO
        # assert (
        #     actual_title == chart_title
        # ), f"Chart title is incorrect. Expected '{chart_title}', got '{actual_title}'."


def test_display_chart(test_file_path):
    """Test if the display_chart function opens the Excel file."""
    create_blank_excel(test_file_path)  # Ensure the file exists
    with mock.patch("xlwings.App") as mock_app:
        mock_instance = mock_app.return_value
        mock_instance.books.open.return_value = mock.MagicMock()
        display_chart(test_file_path)

        # Verify if the Excel file was opened
        mock_instance.books.open.assert_called_once_with(test_file_path)

        # Verify if the application was made visible
        mock_instance.visible = True


def test_main():
    """Test the main function of dixon_graph_generator_exp.py."""

    with mock.patch(
        "dixon_graph_generator_exp.create_blank_excel"
    ) as mock_create, mock.patch(
        "dixon_graph_generator_exp.populate_data"
    ) as mock_populate, mock.patch(
        "dixon_graph_generator_exp.add_chart"
    ) as mock_add_chart, mock.patch(
        "dixon_graph_generator_exp.display_chart"
    ) as mock_display:

        # Call the main function directly
        dixon_graph_generator_exp.main()

        # Assert that all critical functions were called once
        mock_create.assert_called_once()
        mock_populate.assert_called_once()
        mock_add_chart.assert_called_once()
        mock_display.assert_called_once()
